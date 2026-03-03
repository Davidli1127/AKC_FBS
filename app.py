from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
import json
import os
import io
import base64
import copy
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import uuid
from functools import wraps
from dotenv import load_dotenv
load_dotenv()
import db
try:
    import qrcode
    from qrcode.image.pil import PilImage
    QR_AVAILABLE = True
except ImportError:
    QR_AVAILABLE = False
    print('Warning: qrcode library not installed. Run: pip install qrcode[pil]')

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'akc-feedback-secret-key-2026')

ADMIN_ACCOUNT = os.environ.get('ADMIN_ACCOUNT', 'admin')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'akc2026')

def login_required(f):
    """Decorator to require login for admin routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def api_login_required(f):
    """Decorator to require login for API routes (returns JSON)"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return jsonify({'error': 'Authentication required'}), 401
        return f(*args, **kwargs)
    return decorated_function

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR, 'config.json')
DATA_DIR = os.path.join(BASE_DIR, 'data')

# QR code expiration time
COURSE_EXPIRY_HOURS = 24 # Default 24 hours, can adjust if needed

os.makedirs(DATA_DIR, exist_ok=True)

SUBMISSIONS_FILE = os.path.join(DATA_DIR, 'submissions.json')
ALERTS_FILE = os.path.join(DATA_DIR, 'low_feedback_alerts.json')


def load_alerts():
    """Load low feedback alerts from JSON file"""
    if os.path.exists(ALERTS_FILE):
        with open(ALERTS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []


def save_alerts_data(alerts):
    """Save low feedback alerts to JSON file"""
    with open(ALERTS_FILE, 'w', encoding='utf-8') as f:
        json.dump(alerts, f, indent=2, ensure_ascii=False)


def save_low_feedback_alerts(form_id, course_id, course, data, form_config):
    """Detect ratings of 1 or 2 and create alert records"""
    alerts = load_alerts()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    rating_labels = {1: 'Poor', 2: 'Unsatisfactory'}
    q_texts = {}
    for section in form_config.get('sections', []):
        for q in section.get('questions', []):
            q_texts[q['id']] = q['text']

    for key, value in data.items():
        if key.endswith('_comment'):
            continue
        try:
            rating = int(value)
        except (ValueError, TypeError):
            continue
        if rating <= 2:
            q_text = q_texts.get(key, '')
            if not q_text:
                short_id = key.split('_')[-1]
                q_text = q_texts.get(short_id, key)
            comment = data.get(f'{key}_comment', '')
            alert = {
                'id': str(uuid.uuid4())[:12],
                'course_id': course_id,
                'course_title': course.get('course_title', ''),
                'course_date': course.get('course_date', ''),
                'form_id': form_id,
                'participant_name': data.get('name', 'Anonymous'),
                'question_id': key,
                'question_text': q_text,
                'rating': rating,
                'rating_label': rating_labels.get(rating, str(rating)),
                'comment': comment,
                'status': 'new',
                'action_notes': '',
                'submitted_at': now,
                'updated_at': ''
            }
            alerts.append(alert)

    save_alerts_data(alerts)


def load_submissions():
    """Load submission records from JSON file"""
    if os.path.exists(SUBMISSIONS_FILE):
        with open(SUBMISSIONS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def save_submissions_data(submissions):
    """Save submission records to JSON file"""
    with open(SUBMISSIONS_FILE, 'w', encoding='utf-8') as f:
        json.dump(submissions, f, indent=2, ensure_ascii=False)


def _norm_name(name):
    """Normalize a name for comparison: collapse whitespace, uppercase"""
    return ' '.join(name.strip().split()).upper()


def _norm_id(id_number):
    """Normalize an identification number for comparison: strip, uppercase"""
    return id_number.strip().upper()


def has_submitted(course_id, identifier):
    """Check if a student (by ID number) has already submitted for this course"""
    submissions = load_submissions()
    id_norm = _norm_id(identifier)
    return any(_norm_id(s.get('id_number', s.get('name', ''))) == id_norm
               for s in submissions.get(course_id, []))


def record_submission(course_id, participant_name, id_number=''):
    """Record a student submission to prevent duplicates"""
    submissions = load_submissions()
    if course_id not in submissions:
        submissions[course_id] = []
    submissions[course_id].append({
        'name': ' '.join(participant_name.strip().split()),
        'id_number': id_number.strip().upper(),
        'submitted_at': datetime.now().isoformat()
    })
    save_submissions_data(submissions)


def load_config():
    """Load configuration from JSON file"""
    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)


def save_config(config):
    """Save configuration to JSON file"""
    with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=2, ensure_ascii=False)


def clean_expired_courses():
    """Remove courses that are older than COURSE_EXPIRY_HOURS (default 24 hours)"""
    config = load_config()
    now = datetime.now()
    
    original_count = len(config['courses'])
    valid_courses = []
    
    for course in config['courses']:
        created_at = course.get('created_at', '')
        if created_at:
            try:
                created_time = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')
                age_hours = (now - created_time).total_seconds() / 3600
                
                if age_hours < COURSE_EXPIRY_HOURS:
                    valid_courses.append(course)
                else:
                    print(f"Removing expired course: {course.get('course_title', 'Unknown')} (created {age_hours:.1f} hours ago)")
            except ValueError:
                valid_courses.append(course)
        else:
            valid_courses.append(course)
    
    removed_count = original_count - len(valid_courses)
    
    if removed_count > 0:
        config['courses'] = valid_courses
        save_config(config)
        print(f"Cleaned up {removed_count} expired course(s)")
    
    return removed_count

FORM_FILE_NAMES = {
    'form1': 'Trainer_Evaluation',
    'form2': 'Assessor_Evaluation'
}


def get_excel_path(form_id):
    """Get the Excel file path for a form"""
    file_name = FORM_FILE_NAMES.get(form_id, form_id)
    return os.path.join(DATA_DIR, f'{file_name}_responses.xlsx')


def init_excel(form_id):
    """Initialize Excel file with headers if it doesn't exist"""
    excel_path = get_excel_path(form_id)
    
    if os.path.exists(excel_path):
        return
    
    config = load_config()
    form = config['forms'].get(form_id)
    if not form:
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Responses"
    
    if form_id == 'form1':
        headers = [
            'COURSE NAME',
            'DATE',
            'CLASSROOM',
            'INSTRUCTOR 1',
            'INSTRUCTOR 2',
            'INSTRUCTOR 3',
            'LANGUAGE'
        ]
        
        for section in form['sections']:
            if section['type'] == 'rating':
                for q in section['questions']:
                    headers.append(f"{q['id']} - {q['text']}")
            elif section['type'] == 'instructor_rating':
                for q in section['questions']:
                    headers.append(f"B1{q['id']} - {q['text']}")
                for inst_num in range(2, 4):
                    for q in section['questions']:
                        headers.append(f"B{inst_num}{q['id']}")
            elif section['type'] == 'text_questions':
                for q in section['questions']:
                    headers.append(f"{q['id']} - {q['text']}")
    elif form_id == 'form2':
        headers = [
            'COURSE NAME',
            'DATE',
            'CLASSROOM',
            'ASSESSOR 1',
            'ASSESSOR 2',
            'LANGUAGE'
        ]
        
        for section in form['sections']:
            if section['type'] == 'assessor_rating':
                for q in section['questions']:
                    headers.append(f"A1.{q['id']} - {q['text']}")
                for q in section['questions']:
                    headers.append(f"A2.{q['id']}")
            elif section['type'] == 'text_questions':
                for q in section['questions']:
                    headers.append(f"{q['id']} - {q['text']}")
    else:
        headers = ['Submission Time']
        
        for field in form['headerFields']:
            headers.append(field['label'].replace(' (Optional)', ''))
        
        for section in form['sections']:
            if section['type'] == 'rating':
                for q in section['questions']:
                    headers.append(q['id'])
            elif section['type'] == 'instructor_rating':
                # Up to 3 instructors
                for i in range(1, 4):
                    headers.append(f'Instructor {i} Name')
                    for q in section['questions']:
                        headers.append(f'B{i}-{q["id"]}')
            elif section['type'] == 'assessor_rating':
                # Up to 2 assessors
                for i in range(1, 3):
                    headers.append(f'Assessor {i} Name')
                    for q in section['questions']:
                        headers.append(f'A{i}-{q["id"]}')
            elif section['type'] == 'text_questions':
                for q in section['questions']:
                    headers.append(q['id'])
    
    ws.append(headers)
    wb.save(excel_path)

def save_response(form_id, course_id, data):
    """Save form response to Excel and database"""
    excel_path = get_excel_path(form_id)
    init_excel(form_id)
    
    config = load_config()
    form = config['forms'].get(form_id)
    
    course = None
    for c in config['courses']:
        if c['id'] == course_id:
            course = c
            break
    
    # Save to Excel
    wb = load_workbook(excel_path)
    ws = wb.active
    existing_headers = [cell.value for cell in ws[1]]
    data_map = {}
    
    if form_id == 'form1':
        instructors = course.get('instructors', []) if course else []
        data_map['COURSE NAME'] = course['course_title'] if course else ''
        data_map['DATE'] = course['course_date'] if course else ''
        data_map['CLASSROOM'] = course.get('classroom', '') if course else ''
        data_map['INSTRUCTOR 1'] = instructors[0] if len(instructors) > 0 else ''
        data_map['INSTRUCTOR 2'] = instructors[1] if len(instructors) > 1 else ''
        data_map['INSTRUCTOR 3'] = instructors[2] if len(instructors) > 2 else ''
        data_map['LANGUAGE'] = ''
        
        for section in form['sections']:
            if section['type'] == 'rating':
                for q in section['questions']:
                    data_map[q['id']] = data.get(q['id'], '')
            elif section['type'] == 'instructor_rating':
                for inst_num in range(1, 4):
                    for q in section['questions']:
                        key = f"B{inst_num}{q['id']}"
                        data_map[key] = data.get(f'B{inst_num}_{q["id"]}', '')
            elif section['type'] == 'text_questions':
                for q in section['questions']:
                    data_map[q['id']] = data.get(q['id'], '')
                    
    elif form_id == 'form2':
        assessors = course.get('assessors', []) if course else []
        data_map['COURSE NAME'] = course['course_title'] if course else ''
        data_map['DATE'] = course['course_date'] if course else ''
        data_map['CLASSROOM'] = course.get('classroom', '') if course else ''
        data_map['ASSESSOR 1'] = assessors[0] if len(assessors) > 0 else ''
        data_map['ASSESSOR 2'] = assessors[1] if len(assessors) > 1 else ''
        data_map['LANGUAGE'] = ''
        
        for section in form['sections']:
            if section['type'] == 'assessor_rating':
                for i in range(1, 3):
                    for q in section['questions']:
                        key = f"A{i}.{q['id']}"
                        data_map[key] = data.get(f'A{i}_{q["id"]}', '')
            elif section['type'] == 'text_questions':
                for q in section['questions']:
                    data_map[q['id']] = data.get(q['id'], '')
    else:
        data_map['Submission Time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for field in form.get('headerFields', []):
            label = field['label'].replace(' (Optional)', '')
            if field.get('prefilled') and course:
                data_map[label] = course.get(field['id'], '')
            else:
                data_map[label] = data.get(field['id'], '')
        for section in form.get('sections', []):
            if section['type'] in ('rating', 'text_questions'):
                for q in section['questions']:
                    data_map[q['id']] = data.get(q['id'], '')

    row = []
    for header in existing_headers:
        if header is None or header.startswith('[REMOVED]'):
            row.append('')
        else:
            if header in data_map:
                row.append(data_map[header])
            else:
                header_id = header.split(' - ')[0].strip() if ' - ' in header else header
                if header_id in data_map:
                    row.append(data_map[header_id])
                else:
                    row.append('') 
    
    ws.append(row)
    wb.save(excel_path)

    try:
        db.save_to_database(form_id, course_id, course, data)
    except Exception as e:
        print(f"Warning: Could not save to database: {e}")

@app.route('/')
def index():
    """Redirect to admin page"""
    return redirect(url_for('admin'))


@app.route('/scan')
def scan_page():
    """Universal scan / walk-in entry page.
    Participant enters their Class Code and ID number; the server looks up the
    matching active course session and redirects them to the right form.
    """
    return render_template('scan.html')


@app.route('/api/scan/lookup', methods=['POST'])
def scan_lookup():
    """API used by the scan page.
    Accepts { class_code, id_number }.
    Returns the matching course session URL so the client can redirect.
    If the class code matches multiple active sessions (e.g. Form1 + Form2)
    all options are returned and the user can choose.
    """
    data = request.json or {}
    class_code = (data.get('class_code') or '').strip()
    id_number  = (data.get('id_number')  or '').strip()

    if not class_code or not id_number:
        return jsonify({'error': 'Class code and identification number are required.'}), 400

    participant_name = db.get_participant_name_by_id(class_code, id_number)
    if not participant_name:
        return jsonify({'error': 'Identification number not found for that class code. Please check and try again.'}), 404

    config = load_config()
    matches = [c for c in config['courses'] if c['course_title'] == class_code]

    if not matches:
        return jsonify({'error': 'No active QR session found for that class code. Please ask your instructor.'}), 404

    if len(matches) == 1:
        course = matches[0]
        if has_submitted(course['id'], id_number):
            return jsonify({'error': 'You have already submitted feedback for this class. Thank you!'}), 400
        session['student_name'] = participant_name
        session['student_id_number'] = id_number.upper()
        session['student_course_id'] = course['id']
        return jsonify({'redirect': url_for('form_page', course_id=course['id'])})

    options = []
    for c in matches:
        form_label = {
            'form1': 'Trainer Evaluation',
            'form2': 'Assessor Evaluation'
        }.get(c['form_id'], c['form_id'])
        already = has_submitted(c['id'], id_number)
        options.append({
            'course_id': c['id'],
            'form_label': form_label,
            'course_date': c.get('course_date', ''),
            'already_submitted': already
        })
    return jsonify({'options': options, 'participant_name': participant_name,
                    'class_code': class_code, 'id_number': id_number.upper()})


@app.route('/api/scan/select', methods=['POST'])
def scan_select():
    """After user picks one option from multiple sessions."""
    data = request.json or {}
    course_id   = (data.get('course_id')   or '').strip()
    id_number   = (data.get('id_number')   or '').strip()
    participant_name = (data.get('participant_name') or '').strip()

    if not course_id or not id_number:
        return jsonify({'error': 'Missing parameters.'}), 400

    config = load_config()
    course = next((c for c in config['courses'] if c['id'] == course_id), None)
    if not course:
        return jsonify({'error': 'Session not found.'}), 404

    if has_submitted(course_id, id_number):
        return jsonify({'error': 'You have already submitted feedback for this session. Thank you!'}), 400

    session['student_name'] = participant_name
    session['student_id_number'] = id_number.upper()
    session['student_course_id'] = course_id
    return jsonify({'redirect': url_for('form_page', course_id=course_id)})


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Admin login page"""
    if session.get('logged_in'):
        return redirect(url_for('admin'))
    
    error = None
    if request.method == 'POST':
        account = request.form.get('account', '')
        password = request.form.get('password', '')
        
        if account == ADMIN_ACCOUNT and password == ADMIN_PASSWORD:
            session['logged_in'] = True
            session['admin_account'] = account
            return redirect(url_for('admin'))
        else:
            error = 'Invalid account number or password'
    
    return render_template('admin_login.html', error=error)


@app.route('/logout')
def logout():
    """Admin logout"""
    session.clear()
    return redirect(url_for('login'))


@app.route('/admin')
@login_required
def admin():
    """Admin dashboard"""
    clean_expired_courses()
    config = load_config()
    form_personnel = {
        fid: ('assessor' if any(s.get('type') == 'assessor_rating' for s in f.get('sections', [])) else 'instructor')
        for fid, f in config['forms'].items()
    }
    return render_template('admin.html', config=config, form_personnel=form_personnel)

@app.route('/admin/form/<form_id>')
@login_required
def admin_form(form_id):
    """Admin page for editing a specific form"""
    config = load_config()
    form = config['forms'].get(form_id)
    if not form:
        return "Form not found", 404
    return render_template('admin_form.html', form=form, config=config)

@app.route('/api/forms/<form_id>', methods=['GET'])
@api_login_required
def get_form(form_id):
    """Get form configuration"""
    config = load_config()
    form = config['forms'].get(form_id)
    if not form:
        return jsonify({'error': 'Form not found'}), 404
    return jsonify(form)

@app.route('/api/forms/<form_id>', methods=['PUT'])
@api_login_required
def update_form(form_id):
    """Update form configuration"""
    config = load_config()
    if form_id not in config['forms']:
        return jsonify({'error': 'Form not found'}), 404
    
    data = request.json
    config['forms'][form_id] = data
    save_config(config)
    return jsonify({'success': True})

@app.route('/api/forms/<form_id>/sections', methods=['POST'])
@api_login_required
def add_section(form_id):
    """Add a new section to form"""
    config = load_config()
    if form_id not in config['forms']:
        return jsonify({'error': 'Form not found'}), 404
    
    data = request.json
    config['forms'][form_id]['sections'].append(data)
    save_config(config)
    return jsonify({'success': True})

@app.route('/api/forms/<form_id>/sections/<section_id>/questions', methods=['POST'])
def add_question(form_id, section_id):
    """Add a question to a section"""
    config = load_config()
    if form_id not in config['forms']:
        return jsonify({'error': 'Form not found'}), 404
    
    data = request.json
    form = config['forms'][form_id]
    
    for section in form['sections']:
        if section['id'] == section_id:
            section['questions'].append(data)
            break
    
    save_config(config)
    return jsonify({'success': True})

@app.route('/api/forms/<form_id>/sections/<section_id>/questions/<question_id>', methods=['DELETE'])
def delete_question(form_id, section_id, question_id):
    """Delete a question from a section"""
    config = load_config()
    if form_id not in config['forms']:
        return jsonify({'error': 'Form not found'}), 404
    
    form = config['forms'][form_id]
    
    for section in form['sections']:
        if section['id'] == section_id:
            section['questions'] = [q for q in section['questions'] if q['id'] != question_id]
            break
    
    save_config(config)
    return jsonify({'success': True})

@app.route('/api/forms/<form_id>/update-excel', methods=['POST'])
@api_login_required
def update_excel_columns(form_id):
    """Update Excel columns based on form changes"""
    try:
        excel_path = get_excel_path(form_id)
        
        if not os.path.exists(excel_path):
            return jsonify({'success': True, 'message': 'Excel file does not exist yet, will be created on first response'})
        
        changes = request.json
        added_questions = changes.get('addedQuestions', [])
        deleted_questions = changes.get('deletedQuestions', [])
        modified_questions = changes.get('modifiedQuestions', [])
        wb = load_workbook(excel_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        
        for mod in modified_questions:
            q_id = mod['questionId']
            new_text = mod['newText']
            section_id = mod['sectionId']
            for col_idx, header in enumerate(headers):
                if header and header.startswith(f"{q_id} -"):
                    ws.cell(row=1, column=col_idx + 1).value = f"{q_id} - {new_text}"
                    headers[col_idx] = f"{q_id} - {new_text}"
                    break
    
        for added in added_questions:
            q_id = added['questionId']
            q_text = added['questionText']
            section_id = added['sectionId']
            new_header = f"{q_id} - {q_text}"
            new_col = len(headers) + 1
            ws.cell(row=1, column=new_col).value = new_header
            headers.append(new_header)
        
        columns_to_delete = []
        for deleted in deleted_questions:
            q_id = deleted['questionId']
            delete_data = deleted.get('deleteData', False)
            for col_idx, header in enumerate(headers):
                if header and (header.startswith(f"{q_id} -") or header == q_id):
                    if delete_data:
                        columns_to_delete.append(col_idx + 1)  
                    else:
                        ws.cell(row=1, column=col_idx + 1).value = f"[REMOVED] {header}"
                    break
        
        for col_idx in sorted(columns_to_delete, reverse=True):
            ws.delete_cols(col_idx)
        
        wb.save(excel_path)
        
        return jsonify({
            'success': True,
            'message': f'Excel updated: {len(added_questions)} added, {len(modified_questions)} modified, {len(deleted_questions)} deleted'
        })
        
    except Exception as e:
        print(f"Error updating Excel: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/db/test', methods=['GET'])
@api_login_required
def test_db_connection():
    """Test database connection"""
    success, message = db.test_connection()
    return jsonify({'success': success, 'message': message})

@app.route('/api/db/courses', methods=['GET'])
@api_login_required
def search_db_courses():
    """Search courses from database"""
    search_term = request.args.get('search', '')
    limit = request.args.get('limit', 50, type=int)
    courses = db.get_courses_from_db(search_term, limit)
    return jsonify(courses)

@app.route('/api/db/participants', methods=['GET'])
@api_login_required
def get_participants():
    """Get participants for a class with pagination"""
    class_code = request.args.get('class_code', '')
    offset = request.args.get('offset', 0, type=int)
    limit = request.args.get('limit', 20, type=int)
    
    if not class_code:
        return jsonify({'error': 'Class code required'}), 400
    
    result = db.get_participants_by_class(class_code, offset, limit)
    return jsonify(result)

@app.route('/api/db/update-survey-sent', methods=['POST'])
@api_login_required
def update_survey_sent():
    """Update Survey Sent flag for a participant"""
    data = request.json
    course_code = data.get('course_code')
    participant_name = data.get('participant_name')
    
    if not course_code or not participant_name:
        return jsonify({'error': 'Course code and participant name required'}), 400
    
    success = db.update_survey_sent(course_code, participant_name, True)
    return jsonify({'success': success})

@app.route('/api/db/create-tables', methods=['POST'])
@api_login_required
def create_tables():
    """Create feedback tables in database"""
    success, message = db.create_feedback_tables()
    return jsonify({'success': success, 'message': message})

@app.route('/api/db/course-dates', methods=['GET'])
@api_login_required
def get_course_dates():
    """Get all available registration dates from database"""
    dates = db.get_course_dates()
    return jsonify(dates)

@app.route('/api/db/class-codes', methods=['GET'])
@api_login_required
def get_class_codes():
    """Get class codes for a specific registration date"""
    date = request.args.get('date', '')
    if not date:
        return jsonify({'error': 'Date required'}), 400
    codes = db.get_class_codes_by_date(date)
    return jsonify(codes)

@app.route('/api/courses', methods=['GET'])
@api_login_required
def get_courses():
    """Get all courses"""
    config = load_config()
    return jsonify(config['courses'])

@app.route('/api/courses', methods=['POST'])
@api_login_required
def create_course():
    """Create a new course and generate QR code"""
    config = load_config()
    data = request.json
    course = {
        'id': str(uuid.uuid4())[:8],
        'form_id': data['form_id'],
        'course_title': data['course_title'],
        'course_date': data['course_date'],
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

    if data['form_id'] == 'form2':
        course['assessment_location'] = data.get('assessment_location', '')
        course['num_assessors'] = data.get('num_assessors', 1)
        course['assessors'] = data.get('assessors', [])
    else:
        course['classroom'] = data.get('classroom', '')
        course['num_instructors'] = data.get('num_instructors', 1)
        course['instructors'] = data.get('instructors', [])

    config['courses'].append(course)
    save_config(config)

    # Generate QR code
    qr_data = None
    if QR_AVAILABLE:
        public_base = os.environ.get('PUBLIC_URL', '').rstrip('/') or request.host_url.rstrip('/')
        form_url = public_base + f'/student-login/{course["id"]}'
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(form_url)
        qr.make(fit=True)
        img = qr.make_image(fill_color='black', back_color='white')
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        qr_data = base64.b64encode(buf.read()).decode('utf-8')

    return jsonify({'success': True, 'course': course, 'qr_code': qr_data})

@app.route('/api/courses/<course_id>', methods=['DELETE'])
@api_login_required
def delete_course(course_id):
    """Delete a course"""
    config = load_config()
    config['courses'] = [c for c in config['courses'] if c['id'] != course_id]
    save_config(config)
    return jsonify({'success': True})

@app.route('/api/courses/<course_id>/qrcode', methods=['GET'])
@api_login_required
def get_course_qrcode(course_id):
    """Get QR code image for a course"""
    config = load_config()
    course = next((c for c in config['courses'] if c['id'] == course_id), None)
    if not course:
        return jsonify({'error': 'Course not found'}), 404
    if not QR_AVAILABLE:
        return jsonify({'error': 'QR code library not installed'}), 500
    public_base = os.environ.get('PUBLIC_URL', '').rstrip('/') or request.host_url.rstrip('/')
    form_url = public_base + f'/student-login/{course_id}'
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(form_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return send_file(buf, mimetype='image/png', as_attachment=True,
                     download_name=f'QR_{course["course_title"]}_{course["course_date"]}.png')


@app.route('/api/scan/qrcode', methods=['GET'])
@api_login_required
def get_universal_qrcode():
    """Return the universal (fixed) QR code that points to /scan.
    This QR code never changes — participants scan it and then type their
    class code + ID number on the /scan page.
    """
    if not QR_AVAILABLE:
        return jsonify({'error': 'QR code library not installed'}), 500
    public_base = os.environ.get('PUBLIC_URL', '').rstrip('/') or request.host_url.rstrip('/')
    scan_url = public_base + '/scan'
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(scan_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return send_file(buf, mimetype='image/png', as_attachment=True,
                     download_name='AKC_Universal_QR.png')


@app.route('/student-login/<course_id>', methods=['GET', 'POST'])
def student_login(course_id):
    """Student login page - verifies the student by identification number"""
    config = load_config()
    course = next((c for c in config['courses'] if c['id'] == course_id), None)
    if not course:
        return "Course not found. The QR code may be invalid or expired.", 404

    error = None
    if request.method == 'POST':
        id_number = request.form.get('id_number', '').strip()
        if not id_number:
            error = 'Please enter your Identification Number.'
        elif has_submitted(course_id, id_number):
            error = 'You have already submitted feedback for this class. Thank you!'
        else:
            participant_name = db.get_participant_name_by_id(
                course['course_title'], id_number
            )
            if participant_name:
                session['student_name'] = participant_name
                session['student_id_number'] = id_number.upper()
                session['student_course_id'] = course_id
                return redirect(url_for('form_page', course_id=course_id))
            else:
                error = 'Your Identification Number was not found in the participant list for this class. Please check and try again.'
    return render_template('student_login.html', course=course, error=error)

@app.route('/form/<course_id>')
def form_page(course_id):
    """Public form page for participants to fill"""
    # Students must pass through the login page first
    if session.get('student_course_id') != course_id:
        return redirect(url_for('student_login', course_id=course_id))

    config = load_config()
    course = None
    for c in config['courses']:
        if c['id'] == course_id:
            course = c
            break
    
    if not course:
        return "Course not found. The link may be invalid or expired.", 404
    
    form = config['forms'].get(course['form_id'])
    if not form:
        return "Form not found", 404
    
    student_name = session.get('student_name', '')
    return render_template('form.html', form=form, course=course, student_name=student_name)

@app.route('/api/submit/<course_id>', methods=['POST'])
def submit_form(course_id):
    """Submit form response"""
    if session.get('student_course_id') != course_id:
        return jsonify({'error': 'Unauthorized. Please scan the QR code and log in first.'}), 401

    student_name = session.get('student_name', '')
    student_id = session.get('student_id_number', '')

    if has_submitted(course_id, student_id or student_name):
        return jsonify({'error': 'You have already submitted feedback for this class.'}), 400

    config = load_config()
    course = None
    for c in config['courses']:
        if c['id'] == course_id:
            course = c
            break
    
    if not course:
        return jsonify({'error': 'Course not found'}), 404
    
    data = request.json
    save_response(course['form_id'], course_id, data)
    record_submission(course_id, student_name, student_id)
    form_config = config['forms'].get(course['form_id'], {})
    try:
        save_low_feedback_alerts(course['form_id'], course_id, course, data, form_config)
    except Exception as e:
        print(f"Warning: Could not save low feedback alerts: {e}")

    session.pop('student_name', None)
    session.pop('student_id_number', None)
    session.pop('student_course_id', None)

    return jsonify({'success': True, 'message': 'Thank you for your feedback!'})

@app.route('/api/alerts', methods=['GET'])
@api_login_required
def get_alerts():
    """Get low feedback alerts, optionally filtered by status"""
    alerts = load_alerts()
    status_filter = request.args.get('status', '')
    if status_filter:
        alerts = [a for a in alerts if a.get('status') == status_filter]
    alerts.sort(key=lambda a: a.get('submitted_at', ''), reverse=True)
    return jsonify(alerts)

@app.route('/api/alerts/summary', methods=['GET'])
@api_login_required
def get_alerts_summary():
    """Get alert counts grouped by status"""
    alerts = load_alerts()
    summary = {'new': 0, 'acknowledged': 0, 'in_progress': 0, 'resolved': 0, 'total': len(alerts)}
    for a in alerts:
        status = a.get('status', 'new')
        if status in summary:
            summary[status] += 1
    return jsonify(summary)

@app.route('/api/alerts/<alert_id>', methods=['PUT'])
@api_login_required
def update_alert(alert_id):
    """Update alert status and/or action notes"""
    alerts = load_alerts()
    data = request.json
    for alert in alerts:
        if alert['id'] == alert_id:
            if 'status' in data:
                alert['status'] = data['status']
            if 'action_notes' in data:
                alert['action_notes'] = data['action_notes']
            alert['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            break
    else:
        return jsonify({'error': 'Alert not found'}), 404
    save_alerts_data(alerts)
    return jsonify({'success': True})

@app.route('/api/forms', methods=['POST'])
@api_login_required
def create_form():
    """Create a new custom form"""
    config = load_config()
    data = request.json

    form_id = data.get('form_id', '').strip().lower().replace(' ', '_')
    if not form_id:
        return jsonify({'error': 'Form ID is required'}), 400
    if form_id in config['forms']:
        return jsonify({'error': f'Form ID "{form_id}" already exists'}), 400

    template_id = data.get('copy_from', '')
    if template_id and template_id in config['forms']:
        new_form = copy.deepcopy(config['forms'][template_id])
        new_form['id'] = form_id
        new_form['title'] = data.get('title', new_form['title'])
        new_form['formNumber'] = data.get('formNumber', new_form.get('formNumber', ''))
        new_form['description'] = data.get('description', new_form.get('description', ''))
    else:
        new_form = {
            'id': form_id,
            'title': data.get('title', 'New Form'),
            'formNumber': data.get('formNumber', ''),
            'description': data.get('description', ''),
            'headerFields': [
                {'id': 'course_title', 'label': 'Course Title', 'type': 'text', 'required': True, 'prefilled': True},
                {'id': 'course_date', 'label': 'Course Date', 'type': 'date', 'required': True, 'prefilled': True},
                {'id': 'classroom', 'label': 'Classroom', 'type': 'text', 'required': True, 'prefilled': True},
                {'id': 'name', 'label': 'Name', 'type': 'text', 'required': False, 'prefilled': False},
                {'id': 'position', 'label': 'Position', 'type': 'text', 'required': True, 'prefilled': False},
            ],
            'ratingOptions': [
                {'value': 1, 'label': 'Poor'},
                {'value': 2, 'label': 'Unsatisfactory'},
                {'value': 3, 'label': 'Satisfactory'},
                {'value': 4, 'label': 'Very Good'},
                {'value': 5, 'label': 'Excellent'},
            ],
            'sections': data.get('sections', [])
        }

    config['forms'][form_id] = new_form
    save_config(config)
    return jsonify({'success': True, 'form': new_form})

@app.route('/api/forms/<form_id>', methods=['DELETE'])
@api_login_required
def delete_form(form_id):
    """Delete a custom form (built-in form1/form2 are protected)"""
    if form_id in ('form1', 'form2'):
        return jsonify({'error': 'Cannot delete built-in forms'}), 400
    config = load_config()
    if form_id not in config['forms']:
        return jsonify({'error': 'Form not found'}), 404
    del config['forms'][form_id]
    save_config(config)
    return jsonify({'success': True})

if __name__ == '__main__':
    print("Checking for expired course links...")
    clean_expired_courses()
    
    config = load_config()
    for form_id in config['forms']:
        init_excel(form_id)
    
    app.run(debug=True, host='0.0.0.0', port=5000)
